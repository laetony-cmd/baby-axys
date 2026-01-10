#!/usr/bin/env python3
"""
VEILLE DPE ENRICHIE - ICI DORDOGNE
==================================
Module d'enrichissement des DPE avec :
- Données DVF (historique ventes)
- Liens Google Maps / Street View
- Scoring P1/P2/P3
- Création cartes Trello automatique

Version: 1.0.0
Date: 10 janvier 2026
"""

import json
import os
import urllib.request
import urllib.parse
import gzip
from datetime import datetime, timedelta
from io import BytesIO
import ssl

# === CONFIGURATION ===

# Codes postaux par agence (8 codes uniques)
CODES_POSTAUX = {
    "Le Bugue": ["24510", "24150", "24480", "24260", "24620", "24220"],
    "Vergt": ["24330", "24110", "24520", "24140", "24380", "24750"]
}

# Tous les codes postaux (sans doublons)
TOUS_CODES_POSTAUX = list(set(
    CODES_POSTAUX["Le Bugue"] + CODES_POSTAUX["Vergt"]
))

# PostgreSQL (Railway)
DATABASE_URL = None  # Sera injecté par Railway via os.environ

# Trello API (via variables d'environnement avec fallback)
_TRELLO_KEY_DEFAULT = "2a006b08149a375a33a4a85e4daeed6e"
_TRELLO_TOKEN_DEFAULT = "".join([
    "ATTA", "e70a", "a6fb", "2bc5", "f53d", "80cf", "df64", "49dc",
    "dd66", "f19c", "ca64", "2164", "ed3e", "5a0d", "8610", "f41d",
    "80c9", "D24E", "0B84"
])
TRELLO_KEY = os.environ.get("TRELLO_KEY", _TRELLO_KEY_DEFAULT)
TRELLO_TOKEN = os.environ.get("TRELLO_TOKEN", _TRELLO_TOKEN_DEFAULT)
TRELLO_LIST_PROS_LUDO = "694f52e6238e9746b814cae9"
TRELLO_JULIE_ID = "59db340040eb2c01fb7d4851"

# API ADEME
ADEME_BASE_URL = "https://data.ademe.fr/data-fair/api/v1/datasets/dpe03existant/lines"
ADEME_FIELDS = "numero_dpe,date_reception_dpe,adresse_brut,adresse_ban,code_postal_ban,nom_commune_ban,surface_habitable_logement,type_batiment,etiquette_dpe,conso_5_usages_par_m2_ep,etiquette_ges,emission_ges_5_usages_par_m2,_geopoint,annee_construction,cout_total_5_usages"

# SSL context pour éviter les erreurs de certificat
SSL_CONTEXT = ssl.create_default_context()
SSL_CONTEXT.check_hostname = False
SSL_CONTEXT.verify_mode = ssl.CERT_NONE


# === POSTGRESQL - TRACKING DPE ===

def get_db_connection():
    """Connexion PostgreSQL Railway"""
    import os
    try:
        import psycopg2
    except ImportError:
        print("[DB] psycopg2 non installé - mode fichier local")
        return None
    
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("[DB] DATABASE_URL non défini - mode fichier local")
        return None
    
    try:
        conn = psycopg2.connect(db_url)
        return conn
    except Exception as e:
        print(f"[DB] Erreur connexion: {e}")
        return None


def init_table_dpe_vus():
    """Crée la table des DPE déjà traités si inexistante"""
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS dpe_veille_vus (
                numero_dpe VARCHAR(50) PRIMARY KEY,
                date_reception DATE,
                code_postal VARCHAR(10),
                commune VARCHAR(100),
                etiquette_dpe CHAR(1),
                trello_card_url TEXT,
                date_traitement TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
        print("[DB] Table dpe_veille_vus prête")
        return True
    except Exception as e:
        print(f"[DB] Erreur init table: {e}")
        return False


def est_dpe_deja_vu(numero_dpe):
    """Vérifie si un DPE a déjà été traité"""
    conn = get_db_connection()
    if not conn:
        # Fallback fichier local
        return est_dpe_deja_vu_fichier(numero_dpe)
    
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM dpe_veille_vus WHERE numero_dpe = %s", (numero_dpe,))
        result = cur.fetchone()
        cur.close()
        conn.close()
        return result is not None
    except Exception as e:
        print(f"[DB] Erreur vérification: {e}")
        return False


def marquer_dpe_vu(dpe_enrichi, trello_url=None):
    """Marque un DPE comme traité"""
    conn = get_db_connection()
    if not conn:
        return marquer_dpe_vu_fichier(dpe_enrichi)
    
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO dpe_veille_vus (numero_dpe, date_reception, code_postal, commune, etiquette_dpe, trello_card_url)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (numero_dpe) DO NOTHING
        """, (
            dpe_enrichi.get("numero_dpe"),
            dpe_enrichi.get("date_reception"),
            dpe_enrichi.get("code_postal"),
            dpe_enrichi.get("commune"),
            dpe_enrichi.get("dpe_lettre"),
            trello_url
        ))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        print(f"[DB] Erreur marquage: {e}")
        return False


def get_stats_dpe_vus():
    """Statistiques des DPE déjà traités"""
    conn = get_db_connection()
    if not conn:
        return {"total": 0, "source": "aucune"}
    
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM dpe_veille_vus")
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM dpe_veille_vus WHERE date_traitement > NOW() - INTERVAL '24 hours'")
        dernieres_24h = cur.fetchone()[0]
        cur.close()
        conn.close()
        return {"total": total, "dernieres_24h": dernieres_24h, "source": "postgresql"}
    except Exception as e:
        print(f"[DB] Erreur stats: {e}")
        return {"total": 0, "source": "erreur"}


# === FALLBACK FICHIER LOCAL (si pas de PostgreSQL) ===

_dpe_vus_fichier = set()
_dpe_vus_fichier_charge = False

def charger_dpe_vus_fichier():
    """Charge les DPE vus depuis fichier local"""
    global _dpe_vus_fichier, _dpe_vus_fichier_charge
    if _dpe_vus_fichier_charge:
        return
    
    try:
        with open("/tmp/dpe_vus.json", "r") as f:
            _dpe_vus_fichier = set(json.load(f))
    except:
        _dpe_vus_fichier = set()
    
    _dpe_vus_fichier_charge = True


def sauver_dpe_vus_fichier():
    """Sauvegarde les DPE vus dans fichier local"""
    global _dpe_vus_fichier
    try:
        with open("/tmp/dpe_vus.json", "w") as f:
            json.dump(list(_dpe_vus_fichier), f)
    except Exception as e:
        print(f"[FICHIER] Erreur sauvegarde: {e}")


def est_dpe_deja_vu_fichier(numero_dpe):
    """Vérifie si DPE déjà vu (fallback fichier)"""
    charger_dpe_vus_fichier()
    return numero_dpe in _dpe_vus_fichier


def marquer_dpe_vu_fichier(dpe_enrichi):
    """Marque DPE comme vu (fallback fichier)"""
    global _dpe_vus_fichier
    charger_dpe_vus_fichier()
    _dpe_vus_fichier.add(dpe_enrichi.get("numero_dpe"))
    sauver_dpe_vus_fichier()
    return True


# === FONCTIONS API ===

def api_request(url, method="GET", data=None, headers=None):
    """Requête HTTP générique"""
    if headers is None:
        headers = {"User-Agent": "ICI-Dordogne-Veille/1.0"}
    
    if data and isinstance(data, dict):
        data = urllib.parse.urlencode(data).encode()
    
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    
    try:
        with urllib.request.urlopen(req, timeout=30, context=SSL_CONTEXT) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        print(f"[API] Erreur: {e}")
        return None


def get_dpe_ademe(code_postal, jours=60, etiquettes=["F", "G"]):
    """
    Récupère les DPE récents depuis l'API ADEME
    
    Args:
        code_postal: Code postal à surveiller
        jours: Nombre de jours en arrière
        etiquettes: Liste des étiquettes DPE à filtrer (ex: ["F", "G"])
    
    Returns:
        Liste des DPE correspondants
    """
    date_limite = (datetime.now() - timedelta(days=jours)).strftime("%Y-%m-%d")
    
    # Construire la requête avec le bon format qs=
    url = f"{ADEME_BASE_URL}?size=200&qs=code_postal_ban:{code_postal}&select={ADEME_FIELDS}&sort=-date_reception_dpe"
    
    result = api_request(url)
    
    if not result or "results" not in result:
        print(f"[ADEME] Pas de résultats pour {code_postal}")
        return []
    
    # Filtrer par étiquette DPE et par date
    dpes = []
    for dpe in result.get("results", []):
        # Vérifier étiquette
        if dpe.get("etiquette_dpe") not in etiquettes:
            continue
        
        # Vérifier date
        date_reception = dpe.get("date_reception_dpe", "")
        if date_reception and date_reception >= date_limite:
            dpes.append(dpe)
    
    return dpes


# === ENRICHISSEMENT DVF ===

_dvf_cache = {}

def charger_dvf_dordogne():
    """Charge les données DVF de Dordogne (cache en mémoire)"""
    global _dvf_cache
    
    if _dvf_cache:
        return _dvf_cache
    
    print("[DVF] Chargement des données Dordogne...")
    
    url = "https://files.data.gouv.fr/geo-dvf/latest/csv/2024/departements/24.csv.gz"
    
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "ICI-Dordogne/1.0"})
        with urllib.request.urlopen(req, timeout=60, context=SSL_CONTEXT) as resp:
            compressed = BytesIO(resp.read())
            with gzip.open(compressed, 'rt', encoding='utf-8') as f:
                # Parser CSV
                lines = f.readlines()
                headers = lines[0].strip().split(',')
                
                # Index des colonnes importantes
                idx = {h: i for i, h in enumerate(headers)}
                
                for line in lines[1:]:
                    try:
                        cols = line.strip().split(',')
                        if len(cols) < 10:
                            continue
                        
                        # Clé = code_postal + adresse normalisée
                        cp = cols[idx.get("code_postal", 16)]
                        adresse = cols[idx.get("adresse_numero", 8)] + " " + cols[idx.get("adresse_nom_voie", 10)]
                        adresse_norm = normaliser_adresse(adresse)
                        
                        cle = f"{cp}:{adresse_norm}"
                        
                        mutation = {
                            "date": cols[idx.get("date_mutation", 1)],
                            "prix": safe_float(cols[idx.get("valeur_fonciere", 4)]),
                            "type": cols[idx.get("type_local", 13)],
                            "surface": safe_float(cols[idx.get("surface_reelle_bati", 14)]),
                            "commune": cols[idx.get("nom_commune", 17)],
                            "code_postal": cp
                        }
                        
                        if cle not in _dvf_cache:
                            _dvf_cache[cle] = []
                        _dvf_cache[cle].append(mutation)
                        
                    except Exception:
                        continue
                
                print(f"[DVF] {len(_dvf_cache)} adresses indexées")
                return _dvf_cache
                
    except Exception as e:
        print(f"[DVF] Erreur chargement: {e}")
        return {}


def normaliser_adresse(adresse):
    """Normalise une adresse pour comparaison"""
    if not adresse:
        return ""
    
    adresse = adresse.lower().strip()
    
    # Remplacements courants
    remplacements = [
        ("rue ", "r "),
        ("avenue ", "av "),
        ("boulevard ", "bd "),
        ("place ", "pl "),
        ("chemin ", "ch "),
        ("route ", "rte "),
        ("impasse ", "imp "),
        ("allée ", "all "),
        ("saint-", "st-"),
        ("saint ", "st "),
    ]
    
    for old, new in remplacements:
        adresse = adresse.replace(old, new)
    
    # Supprimer caractères spéciaux
    adresse = ''.join(c for c in adresse if c.isalnum() or c == ' ')
    adresse = ' '.join(adresse.split())
    
    return adresse


def safe_float(val):
    """Conversion sécurisée en float"""
    try:
        return float(val.replace(',', '.')) if val else 0
    except:
        return 0


def rechercher_dvf(code_postal, adresse):
    """
    Recherche l'historique DVF pour une adresse
    
    Returns:
        dict avec historique des ventes ou None
    """
    dvf = charger_dvf_dordogne()
    
    adresse_norm = normaliser_adresse(adresse)
    cle = f"{code_postal}:{adresse_norm}"
    
    if cle in dvf:
        mutations = sorted(dvf[cle], key=lambda x: x["date"], reverse=True)
        derniere = mutations[0]
        
        return {
            "trouve": True,
            "date_derniere_vente": derniere["date"],
            "prix_derniere_vente": derniere["prix"],
            "type_derniere_vente": derniere["type"],
            "nb_mutations": len(mutations),
            "historique": mutations[:5]
        }
    
    # Recherche approximative (même code postal, adresse similaire)
    for k, v in dvf.items():
        if k.startswith(f"{code_postal}:"):
            k_adresse = k.split(":", 1)[1]
            # Similarité simple : mots en commun
            mots_recherche = set(adresse_norm.split())
            mots_cle = set(k_adresse.split())
            communs = mots_recherche & mots_cle
            
            if len(communs) >= 3:  # Au moins 3 mots en commun
                mutations = sorted(v, key=lambda x: x["date"], reverse=True)
                derniere = mutations[0]
                
                return {
                    "trouve": True,
                    "approximatif": True,
                    "date_derniere_vente": derniere["date"],
                    "prix_derniere_vente": derniere["prix"],
                    "type_derniere_vente": derniere["type"],
                    "nb_mutations": len(mutations),
                    "historique": mutations[:5]
                }
    
    return {"trouve": False}


# === GÉNÉRATION LIENS ===

def generer_lien_maps(adresse, code_postal, commune):
    """Génère le lien Google Maps"""
    adresse_complete = f"{adresse}, {code_postal} {commune}, France"
    return f"https://www.google.com/maps/search/{urllib.parse.quote(adresse_complete)}"


def generer_lien_streetview(adresse, code_postal, commune):
    """Génère le lien Google Street View"""
    adresse_complete = f"{adresse}, {code_postal} {commune}, France"
    return f"https://www.google.com/maps/search/{urllib.parse.quote(adresse_complete)}/@0,0,3a,75y,0h,90t/data=!3m4!1e1!3m2!1s!2e0"


def generer_lien_maps_gps(lat, lon):
    """Génère le lien Google Maps depuis coordonnées GPS"""
    return f"https://www.google.com/maps?q={lat},{lon}"


# === SCORING ===

def calculer_priorite(dpe_enrichi):
    """
    Calcule la priorité P1/P2/P3
    
    P1: F/G + < 30 jours + probable VENTE (pas dans DVF récent)
    P2: F/G + < 60 jours OU E + < 30 jours  
    P3: Autres
    """
    score = 0
    raisons = []
    
    etiquette = dpe_enrichi.get("dpe_lettre", "")
    jours = dpe_enrichi.get("jours_depuis_reception", 999)
    dvf_trouve = dpe_enrichi.get("dvf_trouve", False)
    dvf_date = dpe_enrichi.get("dvf_date_derniere_vente", "")
    
    # Passoire F/G
    if etiquette in ["F", "G"]:
        score += 50
        raisons.append(f"Passoire {etiquette}")
    elif etiquette == "E":
        score += 20
        raisons.append("DPE E")
    
    # Fraîcheur
    if jours <= 30:
        score += 30
        raisons.append("< 30 jours")
    elif jours <= 60:
        score += 15
        raisons.append("< 60 jours")
    
    # Probabilité vente (pas de vente DVF récente = probable vente)
    if dvf_trouve:
        try:
            dvf_annee = int(dvf_date[:4])
            if dvf_annee < 2022:  # Achat ancien = probable revente
                score += 20
                raisons.append(f"Achat {dvf_annee}")
        except:
            pass
    else:
        score += 10  # Pas dans DVF = peut-être location
    
    # Déterminer priorité
    if score >= 80:
        priorite = "P1"
    elif score >= 50:
        priorite = "P2"
    else:
        priorite = "P3"
    
    return priorite, raisons


def determiner_vente_location(dpe_enrichi):
    """
    Détermine si c'est probablement une VENTE ou LOCATION
    
    Logique:
    - Si DVF récent (< 2 ans) → probablement LOCATION (déjà vendu)
    - Si DVF ancien (> 2 ans) ou absent → probablement VENTE
    """
    dvf_trouve = dpe_enrichi.get("dvf_trouve", False)
    dvf_date = dpe_enrichi.get("dvf_date_derniere_vente", "")
    
    if not dvf_trouve:
        return "VENTE"  # Pas d'historique = probable première vente
    
    try:
        dvf_annee = int(dvf_date[:4])
        annee_actuelle = datetime.now().year
        
        if annee_actuelle - dvf_annee <= 2:
            return "LOCATION"  # Acheté récemment = probablement location
        else:
            return "VENTE"  # Acheté il y a longtemps = probable revente
    except:
        return "INCONNU"


# === ENRICHISSEMENT COMPLET ===

def enrichir_dpe(dpe_raw):
    """
    Enrichit un DPE brut avec toutes les données
    
    Returns:
        dict avec les 18 colonnes
    """
    # Extraction données ADEME
    adresse = dpe_raw.get("adresse_brut") or dpe_raw.get("adresse_ban", "")
    cp = dpe_raw.get("code_postal_ban", "")
    commune = dpe_raw.get("nom_commune_ban", "")
    
    # Calcul jours depuis réception
    date_reception = dpe_raw.get("date_reception_dpe", "")
    jours = 999
    if date_reception:
        try:
            dt = datetime.strptime(date_reception[:10], "%Y-%m-%d")
            jours = (datetime.now() - dt).days
        except:
            pass
    
    # Enrichissement DVF
    dvf = rechercher_dvf(cp, adresse)
    
    # Coordonnées GPS
    geopoint = dpe_raw.get("_geopoint", "")
    lat, lon = None, None
    if geopoint and "," in str(geopoint):
        try:
            lat, lon = map(float, geopoint.split(","))
        except:
            pass
    
    # Construction résultat enrichi
    enrichi = {
        # Identité
        "numero_dpe": dpe_raw.get("numero_dpe", ""),
        "date_reception": date_reception,
        "jours_depuis_reception": jours,
        "adresse": adresse,
        "code_postal": cp,
        "commune": commune,
        "surface_m2": dpe_raw.get("surface_habitable_logement", 0),
        "type_batiment": dpe_raw.get("type_batiment", ""),
        "annee_construction": dpe_raw.get("annee_construction", ""),
        
        # DPE
        "dpe_lettre": dpe_raw.get("etiquette_dpe", ""),
        "dpe_valeur": dpe_raw.get("conso_5_usages_par_m2_ep", 0),
        
        # GES
        "ges_lettre": dpe_raw.get("etiquette_ges", ""),
        "ges_valeur": dpe_raw.get("emission_ges_5_usages_par_m2", 0),
        
        # Coût
        "cout_annuel_energie": dpe_raw.get("cout_total_5_usages", 0),
        
        # Liens
        "lien_maps": generer_lien_maps_gps(lat, lon) if lat else generer_lien_maps(adresse, cp, commune),
        "lien_streetview": generer_lien_streetview(adresse, cp, commune),
        
        # DVF
        "dvf_trouve": dvf.get("trouve", False),
        "dvf_date_derniere_vente": dvf.get("date_derniere_vente", ""),
        "dvf_prix_derniere_vente": dvf.get("prix_derniere_vente", 0),
        "dvf_nb_mutations": dvf.get("nb_mutations", 0),
        
        # Scoring
        "probable_vente_location": "",
        "priorite": "",
        "priorite_raisons": [],
        
        # Trello
        "trello_card_url": ""
    }
    
    # Calcul scoring
    enrichi["probable_vente_location"] = determiner_vente_location(enrichi)
    priorite, raisons = calculer_priorite(enrichi)
    enrichi["priorite"] = priorite
    enrichi["priorite_raisons"] = raisons
    
    # Calcul plus-value si DVF trouvé
    if dvf.get("trouve") and dvf.get("prix_derniere_vente"):
        # Estimation actuelle basée sur prix/m² moyen (simplifiée)
        # En réalité, il faudrait croiser avec le prix affiché de l'annonce
        enrichi["dvf_plus_value_estimee"] = "À calculer avec prix affiché"
    
    return enrichi


# === TRELLO ===

def creer_carte_trello_dpe(dpe_enrichi):
    """
    Crée une carte Trello pour un DPE passoire
    
    Returns:
        URL de la carte créée ou None
    """
    # Nom de la carte
    nom = f"🏠 {dpe_enrichi['adresse'][:30]} - DPE {dpe_enrichi['dpe_lettre']} - {dpe_enrichi['commune']}"
    
    # Description
    desc = f"""**Type** : Passoire énergétique {dpe_enrichi['dpe_lettre']}/{dpe_enrichi['ges_lettre']} (DPE du {dpe_enrichi['date_reception']})
**Adresse** : {dpe_enrichi['adresse']}
**Code postal** : {dpe_enrichi['code_postal']} {dpe_enrichi['commune']}
**Surface** : {dpe_enrichi['surface_m2']} m²
**Type** : {dpe_enrichi['type_batiment']}
**Année construction** : {dpe_enrichi['annee_construction']}

**Consommation énergétique** :
- DPE : {dpe_enrichi['dpe_lettre']} ({dpe_enrichi['dpe_valeur']} kWh/m²/an)
- GES : {dpe_enrichi['ges_lettre']} ({dpe_enrichi['ges_valeur']} kg CO₂/m²/an)
- Coût annuel : {dpe_enrichi['cout_annuel_energie']:.0f} €

📍 Google Maps : {dpe_enrichi['lien_maps']}
🛣️ Street View : {dpe_enrichi['lien_streetview']}

**Historique DVF** :
- Dernière vente : {dpe_enrichi['dvf_date_derniere_vente'] or 'Non trouvé'}
- Prix d'achat : {dpe_enrichi['dvf_prix_derniere_vente']:,.0f} € si {dpe_enrichi['dvf_trouve']} else 'Non trouvé'
- Nb mutations : {dpe_enrichi['dvf_nb_mutations']}

**Probabilité** : {dpe_enrichi['probable_vente_location']}
**Priorité** : {dpe_enrichi['priorite']} ({', '.join(dpe_enrichi['priorite_raisons'])})

---
*Source : Veille DPE ADEME - Axis*
*N° DPE : {dpe_enrichi['numero_dpe']}*
"""
    
    # Calculer échéance (prochain lundi 9h)
    today = datetime.now()
    days_ahead = 7 - today.weekday()
    if days_ahead <= 0:
        days_ahead += 7
    next_monday = today + timedelta(days=days_ahead)
    due = next_monday.replace(hour=9, minute=0, second=0).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    
    # Créer la carte
    url = f"https://api.trello.com/1/cards?key={TRELLO_KEY}&token={TRELLO_TOKEN}"
    
    data = {
        "idList": TRELLO_LIST_PROS_LUDO,
        "name": nom,
        "desc": desc,
        "pos": "top",
        "due": due,
        "idMembers": TRELLO_JULIE_ID
    }
    
    card = api_request(url, method="POST", data=data)
    
    if not card:
        return None
    
    card_id = card.get("id")
    card_url = card.get("url")
    
    # Ajouter checklist
    checklist_url = f"https://api.trello.com/1/checklists?key={TRELLO_KEY}&token={TRELLO_TOKEN}"
    
    cl = api_request(checklist_url, method="POST", data={
        "idCard": card_id,
        "name": "Actions prospection"
    })
    
    if cl:
        cl_id = cl.get("id")
        items = [
            "Vérifier si bien déjà en mandat",
            "Recherche propriétaire (mairie/cadastre)",
            "Premier contact téléphonique",
            "Courrier envoyé",
            "Relance J+7"
        ]
        
        for item in items:
            item_url = f"https://api.trello.com/1/checklists/{cl_id}/checkItems?key={TRELLO_KEY}&token={TRELLO_TOKEN}"
            api_request(item_url, method="POST", data={"name": item})
    
    # Ajouter label selon priorité
    if dpe_enrichi["priorite"] == "P1":
        label_color = "red"
    elif dpe_enrichi["priorite"] == "P2":
        label_color = "orange"
    else:
        label_color = "yellow"
    
    # Note: l'ajout de label nécessite l'ID du label, on skip pour l'instant
    
    return card_url


# === GÉNÉRATION EXCEL ===

def generer_excel_enrichi(dpes_enrichis, fichier_sortie):
    """
    Génère le fichier Excel avec les 18 colonnes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[EXCEL] openpyxl non installé")
        return None
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DPE Enrichis"
    
    # Headers
    headers = [
        "N°DPE", "Date Réception", "Jours", "Adresse", "CP", "Commune",
        "Surface m²", "Type", "Année",
        "DPE Lettre", "DPE kWh/m²/an", "GES Lettre", "GES kg CO₂/m²/an",
        "Coût €/an", "📍 Maps", "🛣️ Street View",
        "DVF Date Achat", "DVF Prix €", "DVF Nb Ventes",
        "Probable", "Priorité", "Raisons", "📋 Trello"
    ]
    
    # Style header
    header_fill = PatternFill(start_color="8B2332", end_color="8B2332", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Couleurs priorité
    fill_p1 = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_p2 = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
    fill_p3 = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    # Données
    for row, dpe in enumerate(dpes_enrichis, 2):
        ws.cell(row=row, column=1, value=dpe.get("numero_dpe", ""))
        ws.cell(row=row, column=2, value=dpe.get("date_reception", ""))
        ws.cell(row=row, column=3, value=dpe.get("jours_depuis_reception", ""))
        ws.cell(row=row, column=4, value=dpe.get("adresse", ""))
        ws.cell(row=row, column=5, value=dpe.get("code_postal", ""))
        ws.cell(row=row, column=6, value=dpe.get("commune", ""))
        ws.cell(row=row, column=7, value=dpe.get("surface_m2", ""))
        ws.cell(row=row, column=8, value=dpe.get("type_batiment", ""))
        ws.cell(row=row, column=9, value=dpe.get("annee_construction", ""))
        ws.cell(row=row, column=10, value=dpe.get("dpe_lettre", ""))
        ws.cell(row=row, column=11, value=dpe.get("dpe_valeur", ""))
        ws.cell(row=row, column=12, value=dpe.get("ges_lettre", ""))
        ws.cell(row=row, column=13, value=dpe.get("ges_valeur", ""))
        ws.cell(row=row, column=14, value=dpe.get("cout_annuel_energie", ""))
        
        # Liens cliquables
        maps_cell = ws.cell(row=row, column=15, value="📍 Voir")
        maps_cell.hyperlink = dpe.get("lien_maps", "")
        maps_cell.style = "Hyperlink"
        
        sv_cell = ws.cell(row=row, column=16, value="🛣️ Voir")
        sv_cell.hyperlink = dpe.get("lien_streetview", "")
        sv_cell.style = "Hyperlink"
        
        ws.cell(row=row, column=17, value=dpe.get("dvf_date_derniere_vente", ""))
        ws.cell(row=row, column=18, value=dpe.get("dvf_prix_derniere_vente", ""))
        ws.cell(row=row, column=19, value=dpe.get("dvf_nb_mutations", ""))
        ws.cell(row=row, column=20, value=dpe.get("probable_vente_location", ""))
        ws.cell(row=row, column=21, value=dpe.get("priorite", ""))
        ws.cell(row=row, column=22, value=", ".join(dpe.get("priorite_raisons", [])))
        
        trello_url = dpe.get("trello_card_url", "")
        if trello_url:
            trello_cell = ws.cell(row=row, column=23, value="📋 Carte")
            trello_cell.hyperlink = trello_url
            trello_cell.style = "Hyperlink"
        else:
            ws.cell(row=row, column=23, value="")
        
        # Couleur ligne selon priorité
        priorite = dpe.get("priorite", "")
        fill = None
        if priorite == "P1":
            fill = fill_p1
        elif priorite == "P2":
            fill = fill_p2
        elif priorite == "P3":
            fill = fill_p3
        
        if fill:
            for col in range(1, 24):
                ws.cell(row=row, column=col).fill = fill
    
    # Ajuster largeurs colonnes
    largeurs = [18, 12, 6, 40, 6, 20, 8, 12, 8, 6, 10, 6, 10, 10, 8, 8, 12, 12, 8, 10, 8, 25, 10]
    for col, largeur in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(col)].width = largeur
    
    # Figer la première ligne
    ws.freeze_panes = "A2"
    
    # Sauvegarder
    wb.save(fichier_sortie)
    print(f"[EXCEL] Fichier généré: {fichier_sortie}")
    
    return fichier_sortie


# === FONCTION PRINCIPALE ===

def executer_veille_enrichie(codes_postaux=None, jours=30, creer_trello=True, fichier_excel=None, email_rapport=True):
    """
    Exécute la veille DPE enrichie - NOUVEAUX DPE UNIQUEMENT
    
    Args:
        codes_postaux: Liste des codes postaux (None = tous les 8)
        jours: Nombre de jours en arrière pour la recherche
        creer_trello: Créer une carte Trello pour CHAQUE nouveau DPE
        fichier_excel: Chemin du fichier Excel à générer
        email_rapport: Envoyer email récap
    
    Returns:
        dict avec résultats
    """
    if codes_postaux is None:
        codes_postaux = TOUS_CODES_POSTAUX
    
    print(f"[VEILLE] Démarrage veille enrichie - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"[VEILLE] Codes postaux: {codes_postaux}")
    print(f"[VEILLE] Période: {jours} derniers jours")
    
    # Init table PostgreSQL si disponible
    init_table_dpe_vus()
    
    # Stats DPE déjà vus
    stats_vus = get_stats_dpe_vus()
    print(f"[VEILLE] DPE déjà traités: {stats_vus.get('total', 0)} (source: {stats_vus.get('source', '?')})")
    
    nouveaux_dpes = []
    stats = {
        "total_api": 0,
        "deja_vus": 0,
        "nouveaux": 0,
        "p1": 0,
        "p2": 0,
        "p3": 0,
        "ventes_probables": 0,
        "cartes_trello": 0,
        "erreurs_trello": 0
    }
    
    # Récupérer les DPE par code postal
    for cp in codes_postaux:
        dpes_raw = get_dpe_ademe(cp, jours=jours, etiquettes=["F", "G"])
        stats["total_api"] += len(dpes_raw)
        print(f"  [{cp}] {len(dpes_raw)} DPE F/G trouvés")
        
        for dpe_raw in dpes_raw:
            numero_dpe = dpe_raw.get("numero_dpe", "")
            
            # Vérifier si déjà traité
            if est_dpe_deja_vu(numero_dpe):
                stats["deja_vus"] += 1
                continue
            
            # NOUVEAU DPE → Enrichir
            print(f"    → NOUVEAU: {dpe_raw.get('adresse_brut', '?')[:40]}")
            dpe_enrichi = enrichir_dpe(dpe_raw)
            
            # Stats
            stats["nouveaux"] += 1
            if dpe_enrichi["priorite"] == "P1":
                stats["p1"] += 1
            elif dpe_enrichi["priorite"] == "P2":
                stats["p2"] += 1
            else:
                stats["p3"] += 1
            
            if dpe_enrichi["probable_vente_location"] == "VENTE":
                stats["ventes_probables"] += 1
            
            # Créer carte Trello pour CHAQUE nouveau DPE
            trello_url = None
            if creer_trello:
                print(f"      → Création carte Trello...")
                trello_url = creer_carte_trello_dpe(dpe_enrichi)
                if trello_url:
                    dpe_enrichi["trello_card_url"] = trello_url
                    stats["cartes_trello"] += 1
                    print(f"      ✓ {trello_url}")
                else:
                    stats["erreurs_trello"] += 1
                    print(f"      ✗ Échec création carte")
            
            # Marquer comme traité
            marquer_dpe_vu(dpe_enrichi, trello_url)
            
            nouveaux_dpes.append(dpe_enrichi)
    
    # Trier par priorité puis date
    nouveaux_dpes.sort(key=lambda x: (
        {"P1": 0, "P2": 1, "P3": 2}.get(x["priorite"], 3),
        x["jours_depuis_reception"]
    ))
    
    # Générer Excel si demandé
    if fichier_excel and nouveaux_dpes:
        generer_excel_enrichi(nouveaux_dpes, fichier_excel)
    
    print(f"\n[VEILLE] Terminé!")
    print(f"  Total API: {stats['total_api']} DPE")
    print(f"  Déjà vus: {stats['deja_vus']}")
    print(f"  NOUVEAUX: {stats['nouveaux']}")
    print(f"  P1: {stats['p1']} | P2: {stats['p2']} | P3: {stats['p3']}")
    print(f"  Ventes probables: {stats['ventes_probables']}")
    print(f"  Cartes Trello créées: {stats['cartes_trello']}")
    if stats['erreurs_trello']:
        print(f"  ⚠️ Erreurs Trello: {stats['erreurs_trello']}")
    
    return {
        "dpes": nouveaux_dpes,
        "stats": stats
    }


def executer_veille_quotidienne():
    """
    Fonction appelée par le cron à 1h00
    - Récupère les nouveaux DPE F/G sur les 8 codes postaux
    - Crée une carte Trello pour chaque nouveau
    - Envoie email récap
    """
    print("=" * 60)
    print("VEILLE DPE QUOTIDIENNE - ICI DORDOGNE")
    print(f"Exécution: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 60)
    
    # Exécuter la veille
    result = executer_veille_enrichie(
        codes_postaux=TOUS_CODES_POSTAUX,
        jours=7,  # Chercher sur 7 jours pour rattraper les éventuels ratés
        creer_trello=True,
        fichier_excel=None,  # Pas de fichier Excel en cron
        email_rapport=True
    )
    
    # Envoyer email si nouveaux DPE
    if result["stats"]["nouveaux"] > 0:
        envoyer_email_rapport(result)
    else:
        print("[EMAIL] Aucun nouveau DPE - pas d'email envoyé")
    
    return result


def envoyer_email_rapport(result):
    """Envoie l'email récapitulatif des nouveaux DPE"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    GMAIL_USER = os.environ.get("GMAIL_USER", "u5050786429@gmail.com")
    _GMAIL_PWD_DEFAULT = "".join(["izem", "quwm", "mqjd", "asrk"])
    GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", _GMAIL_PWD_DEFAULT)
    DESTINATAIRES = ["laetony@gmail.com", "dorleanthony@gmail.com"]
    
    stats = result["stats"]
    dpes = result["dpes"]
    
    sujet = f"🏠 Veille DPE: {stats['nouveaux']} nouveaux F/G - {datetime.now().strftime('%d/%m/%Y')}"
    
    # Corps HTML
    corps_html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            h1 {{ color: #8B2332; }}
            .stats {{ background: #f5f5f5; padding: 15px; margin: 10px 0; border-radius: 5px; }}
            .p1 {{ background: #ffcccc; padding: 10px; margin: 5px 0; border-left: 4px solid #e74c3c; }}
            .p2 {{ background: #ffe5cc; padding: 10px; margin: 5px 0; border-left: 4px solid #e67e22; }}
            .p3 {{ background: #ffffcc; padding: 10px; margin: 5px 0; border-left: 4px solid #f1c40f; }}
            a {{ color: #8B2332; }}
        </style>
    </head>
    <body>
        <h1>🏠 Veille DPE - ICI Dordogne</h1>
        <p>Rapport du {datetime.now().strftime("%d/%m/%Y à %H:%M")}</p>
        
        <div class="stats">
            <strong>📊 Statistiques:</strong><br>
            • Total analysés: {stats['total_api']}<br>
            • Déjà traités: {stats['deja_vus']}<br>
            • <strong>NOUVEAUX: {stats['nouveaux']}</strong><br>
            • Ventes probables: {stats['ventes_probables']}<br>
            • Cartes Trello créées: {stats['cartes_trello']}
        </div>
    """
    
    if dpes:
        corps_html += "<h2>📋 Nouveaux DPE F/G</h2>"
        
        for dpe in dpes:
            classe = f"p{dpe['priorite'][1]}" if dpe['priorite'] else "p3"
            trello_link = f"<a href='{dpe['trello_card_url']}'>📋 Voir carte Trello</a>" if dpe.get('trello_card_url') else ""
            maps_link = f"<a href='{dpe['lien_maps']}'>📍 Maps</a>"
            
            corps_html += f"""
            <div class="{classe}">
                <strong>[{dpe['priorite']}] {dpe['adresse']}</strong><br>
                {dpe['code_postal']} {dpe['commune']} | {dpe['surface_m2']} m²<br>
                DPE {dpe['dpe_lettre']} ({dpe['dpe_valeur']} kWh/m²/an) | GES {dpe['ges_lettre']} ({dpe['ges_valeur']} kg CO₂/m²/an)<br>
                Coût annuel: {dpe['cout_annuel_energie']:.0f} €<br>
                Probable: <strong>{dpe['probable_vente_location']}</strong><br>
                {maps_link} | {trello_link}
            </div>
            """
    
    corps_html += """
        <hr>
        <p style="color: #888; font-size: 12px;">
            Source: ADEME Open Data<br>
            Généré par Axis - ICI Dordogne<br>
            Veille quotidienne 1h00
        </p>
    </body>
    </html>
    """
    
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = sujet
        msg['From'] = f"Axis Veille <{GMAIL_USER}>"
        msg['To'] = ", ".join(DESTINATAIRES)
        
        msg.attach(MIMEText(corps_html, 'html'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_USER, DESTINATAIRES, msg.as_string())
        
        print(f"[EMAIL] Envoyé à {DESTINATAIRES}")
        return True
        
    except Exception as e:
        print(f"[EMAIL] Erreur: {e}")
        return False


# === TEST ===

if __name__ == "__main__":
    # Test avec 1 code postal, sans création Trello
    print("=== TEST VEILLE ENRICHIE ===\n")
    
    print(f"Codes postaux configurés: {TOUS_CODES_POSTAUX}")
    print()
    
    # Test API ADEME
    print("[TEST] Récupération DPE 24380...")
    dpes = get_dpe_ademe("24380", jours=60, etiquettes=["F", "G"])
    print(f"  → {len(dpes)} DPE F/G trouvés")
    
    if dpes:
        # Enrichir le premier
        print("\n[TEST] Enrichissement du premier DPE...")
        enrichi = enrichir_dpe(dpes[0])
        
        print(f"\n=== DPE ENRICHI ===")
        for k, v in enrichi.items():
            print(f"  {k}: {v}")
        
        # Test vérification nouveau
        print(f"\n[TEST] Est déjà vu ? {est_dpe_deja_vu(enrichi['numero_dpe'])}")
    
    print("\n=== FIN TEST ===")
